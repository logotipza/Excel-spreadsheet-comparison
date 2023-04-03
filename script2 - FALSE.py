import tkinter as tk
from tkinter import filedialog
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

def select_files_and_columns():
    file_path1 = filedialog.askopenfilename(title="Выберите первую таблицу")
    file_path2 = filedialog.askopenfilename(title="Выберите вторую таблицу")
    output_file_path = filedialog.asksaveasfilename(title="Сохранить результаты", defaultextension=".xlsx")

    col_name1 = column1_entry.get()
    col_name2 = column2_entry.get()

    df1 = pd.read_excel(file_path1)
    df2 = pd.read_excel(file_path2)

    merged_table = pd.concat([df1, df2], ignore_index=True, sort=False, axis=1)
    merged_table['match'] = False

    matched_rows = []
    for i, row1 in df1.iterrows():
        for j, row2 in df2.iterrows():
            if str(row1[col_name1]) in str(row2[col_name2]):
                matched_rows.append((i, j))
                merged_table.loc[i, 'match'] = True
                merged_table.loc[len(df1) + j, 'match'] = True
                break

    wb = Workbook()
    ws = wb.active
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    for row_index, row in enumerate(dataframe_to_rows(merged_table, index=False, header=True)):
        if row_index == 0:  # Заголовок таблицы
            ws.append(row)
        else:
            ws.append(row)
            if not row[-1]:  # Если значение в столбце 'match' равно False
                for cell in ws[row_index + 1]:
                    cell.fill = red_fill

    wb.save(output_file_path)

# Инициализация графического интерфейса
root = tk.Tk()
root.title('Сравнение таблиц')

column1_label = tk.Label(root, text="Столбец в первой таблице:")
column1_label.pack()
column1_entry = tk.Entry(root)
column1_entry.pack()

column2_label = tk.Label(root, text="Столбец во второй таблице:")
column2_label.pack()
column2_entry = tk.Entry(root)
column2_entry.pack()

select_button = tk.Button(root, text='Выбрать файлы и столбцы', command=select_files_and_columns)
select_button.pack()

root.mainloop()
