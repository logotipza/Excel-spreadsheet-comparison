import tkinter as tk
from tkinter import filedialog, ttk
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

def load_columns(file_path, combobox):
    df = pd.read_excel(file_path)
    columns = list(df.columns)
    combobox["values"] = columns
    combobox.current(0)

def select_files_and_columns():
    file_path1 = filedialog.askopenfilename(title="Выберите первую таблицу")
    file_path2 = filedialog.askopenfilename(title="Выберите вторую таблицу")
    output_file_path = filedialog.asksaveasfilename(title="Сохранить результаты", defaultextension=".xlsx")

    def on_file1_selected(*args):
        load_columns(file_path1, column1_combobox)

    def on_file2_selected(*args):
        load_columns(file_path2, column2_combobox)

    column1_label = tk.Label(root, text="Столбец в первой таблице:")
    column1_label.pack()
    column1_combobox = ttk.Combobox(root, postcommand=on_file1_selected)
    column1_combobox.pack()

    column2_label = tk.Label(root, text="Столбец во второй таблице:")
    column2_label.pack()
    column2_combobox = ttk.Combobox(root, postcommand=on_file2_selected)
    column2_combobox.pack()

    def process_tables():
        col_name1 = column1_combobox.get()
        col_name2 = column2_combobox.get()

        df1 = pd.read_excel(file_path1)
        df2 = pd.read_excel(file_path2)

        result_rows = []
        unmatched_rows = []

        for i, row1 in df1.iterrows():
            cell_value1 = row1[col_name1]
            matched = False
            repeated = False
            for j, row2 in df2.iterrows():
                if str(cell_value1) in str(row2[col_name2]):
                    if not matched:
                        matched = True
                    else:
                        repeated = True
                    new_row = row1.tolist() + row2.tolist() + [matched, repeated]
                    result_rows.append(new_row)
            if not matched:
                new_row = row1.tolist() + [None] * len(df2.columns) + [matched, repeated]
                unmatched_rows.append(new_row)

        result_rows.extend(unmatched_rows)
        result_df = pd.DataFrame(result_rows, columns=df1.columns.tolist() + df2.columns.tolist() + ['matched', 'repeated'])

        wb = Workbook()
        ws = wb.active
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

        for row_index, row in enumerate(dataframe_to_rows(result_df, index=False, header=True)):
            if row_index == 0:  # Заголовок таблицы
                ws.append(row)
            else:
                ws.append(row[:-2])
                if not row[-2]:  # Если значение в столбце 'matched' равно False
                    for cell in ws[row_index + 1]:
                        cell.fill = red_fill
                elif row[-1]:  # Если значение в столбце 'repeated' равно True
                    for cell in ws[row_index + 1]:
                        cell.fill = blue_fill

        wb.save(output_file_path)

    process_button = tk.Button(root, text='Обработать таблицы', command=process_tables)
    process_button.pack()

# Инициализация графического интерфейса
root = tk.Tk()
root.title('Сравнение таблиц')

select_button = tk.Button(root, text='Выбрать файлы и столбцы', command=select_files_and_columns)
select_button.pack()

root.mainloop()

               
