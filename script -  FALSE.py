import tkinter as tk
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Функция для выбора файлов, столбцов и строк
def select_files_and_columns():
    file_path1 = filedialog.askopenfilename(title='Выберите первую таблицу')
    file_path2 = filedialog.askopenfilename(title='Выберите вторую таблицу')

    first_table = pd.read_excel(file_path1)
    second_table = pd.read_excel(file_path2)

    first_column = choose_column(first_table.columns, 'Выберите столбец для первой таблицы')
    second_column = choose_column(second_table.columns, 'Выберите столбец для второй таблицы')

    first_row = choose_row(first_table.index, 'Выберите строку для первой таблицы')
    second_row = choose_row(second_table.index, 'Выберите строку для второй таблицы')

    process_tables(file_path1, file_path2, first_column, second_column, first_row, second_row)

# Функция для выбора столбца
def choose_column(columns, title):
    column_window = tk.Toplevel(root)
    column_window.title(title)
    column_var = tk.StringVar(column_window)
    column_var.set(columns[0])
    column_dropdown = tk.OptionMenu(column_window, column_var, *columns)
    column_dropdown.pack()
    confirm_button = tk.Button(column_window, text='Подтвердить', command=column_window.destroy)
    confirm_button.pack()
    column_window.wait_window()
    return column_var.get()

# Функция для выбора строки
def choose_row(rows, title):
    row_window = tk.Toplevel(root)
    row_window.title(title)
    row_var = tk.IntVar(row_window)
    row_var.set(rows[0])
    row_slider = tk.Scale(row_window, from_=rows[0], to=rows[-1], variable=row_var, orient=tk.HORIZONTAL)
    row_slider.pack()
    confirm_button = tk.Button(row_window, text='Подтвердить', command=row_window.destroy)
    confirm_button.pack()
    row_window.wait_window()
    return row_var.get()

# Обработка таблиц и сохранение результата
def process_tables(file_path1, file_path2, first_column, second_column, first_row, second_row):
    first_table = pd.read_excel(file_path1)
    second_table = pd.read_excel(file_path2)

    first_table = first_table.iloc[[first_row]]
    second_table = second_table.iloc[[second_row]]

    merged_table = first_table.merge(second_table, left_on=first_column, right_on=second_column, how='outer')
    merged_table['match'] = merged_table.apply(lambda row: not pd.isnull(row[first_column]) and not pd.isnull(row[second_column]), axis=1)

    output_file_path = filedialog.asksaveasfilename(defaultextension='.xlsx', title='Сохранить результат как')
    merged_table.to_excel(output_file_path, index=False)

    # Окрашивание строк без соответствия
    wb = load_workbook(output_file_path)
    ws = wb.active
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    for row_index, row in enumerate(dataframe_to_rows(merged_table, index=False, header=True)):
        if row_index == 0:  # Заголовок таблицы
            ws.append(row)
        else:
            ws.append(row)
            if not row[-1]:  # Если значение в столбце 'match' равно False
                for cell in ws[row_index + 1]:
                    cell.fill = red_fill

    wb.save(output_file_path)

# Инициализация графического интерфейса
root = tk.Tk()
root.title('Сравнение таблиц')

select_button = tk.Button(root, text='Выбрать файлы и столбцы', command=select_files_and_columns)
select_button.pack()

root.mainloop()

